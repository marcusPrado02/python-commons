"""Application export – ExcelExporter (requires openpyxl extra)."""
from __future__ import annotations

import io
from typing import Any

from mp_commons.application.export.request import ExportRequest

__all__ = ["ExcelExporter"]


def _require_openpyxl() -> Any:  # pragma: no cover
    try:
        import openpyxl  # noqa: PLC0415
        return openpyxl
    except ImportError as exc:
        raise ImportError(
            "openpyxl is required for Excel export. "
            "Install it with: pip install openpyxl"
        ) from exc


class ExcelExporter:
    """Exports data to an .xlsx workbook using ``openpyxl``."""

    async def export(self, request: ExportRequest) -> bytes:  # pragma: no cover
        openpyxl = _require_openpyxl()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = request.filename[:31]  # sheet name limit

        from openpyxl.styles import Font  # noqa: PLC0415

        # Write header row with bold font
        for col_idx, col_def in enumerate(request.columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_def.header)
            cell.font = Font(bold=True)

        row_idx = 2
        async for row in request.rows:
            for col_idx, col_def in enumerate(request.columns, start=1):
                value = row.get(col_def.key, "")
                ws.cell(row=row_idx, column=col_idx, value=value)
            row_idx += 1

        # Auto-size columns
        for col_cells in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col_cells)
            ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 2, 50)

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()
